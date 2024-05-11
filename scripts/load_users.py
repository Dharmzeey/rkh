from openpyxl import load_workbook
from user.models import User


# def run():
#     wb = load_workbook('kutites.xlsx')
#     ws = wb.active

#     m_row = ws.max_row
#     m_col = ws.max_column

#     for row in range(2, m_row):
#         ln = ws.cell(row=row, column=1).value
#         fn = ws.cell(row=row, column=2).value
#         on = ws.cell(row=row, column=3).value
#         dept = ws.cell(row=row, column=4).value
#         email = ws.cell(row=row, column=5).value    
#         user = User.objects.create(
#             last_name = ln,
#             first_name = fn,
#             other_name = on,
#             department = dept,
#             email = email,
#         )
#         user.set_password('w1234')
#         user.save()
        
def run():
    wb = load_workbook('electoral.xlsx')
    ws = wb.active

    m_row = ws.max_row
    m_col = ws.max_column

    for row in range(2, m_row):
        ln = ws.cell(row=row, column=1).value
        fn = ws.cell(row=row, column=2).value
        on = ws.cell(row=row, column=3).value
        dept = ws.cell(row=row, column=4).value
        email = ws.cell(row=row, column=5).value    
        user = User.objects.create(
            last_name = ln,
            first_name = fn,
            other_name = on,
            department = dept,
            email = email,
        )
        user.set_password('w1234')
        user.save()
        